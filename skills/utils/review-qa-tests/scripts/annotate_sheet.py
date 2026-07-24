# -*- coding: utf-8 -*-
"""Annotate an .xlsx with THREADED comments (not legacy notes) + highlighted new rows.

Usage:
    python annotate_sheet.py config.json          # write
    python annotate_sheet.py --verify out.xlsx    # validate a written file

config.json:
    source        .xlsx to read (alias: "backup")
    out           destination .xlsx (defaults to source; a .PREV.xlsx backup is auto-made)
    sheet         worksheet name to annotate
    author        comment author display name
    template_row  int; row whose per-column style new rows copy
    comments      { "G3": "text", ... }  cells on existing rows
    new_rows      [ { "C": "...", "D": "...", "G": "..." }, ... ]  appended after last row
    dt            optional ISO datetime for the comments (default fixed)

Two modes, auto-selected by whether `source` already holds comments:
  - REBUILD   (source has no comment parts): openpyxl round-trip, then convert to threaded.
  - IN-PLACE  (source already has comments — e.g. another sheet was reviewed earlier):
              pure zip surgery on the target sheet, every existing part left byte-identical,
              so threaded comments elsewhere in the workbook are never flattened to notes.

Both write real threaded comments and dodge the two orphan-relationship faults that make
Excel prompt to repair (see EXCEL.md).
"""
import sys, os, re, json, uuid, zipfile, shutil
from xml.sax.saxutils import escape

PERSON_ID = '{0C5AF367-69BB-4C84-BB7C-B41C21C97D64}'
BOILER = ('[Threaded comment]\n\nYour version of Excel allows you to read this threaded '
          'comment; however, any edits to it will get removed if the file is opened in a '
          'newer version of Excel. Learn more: https://go.microsoft.com/fwlink/?linkid=870924'
          '\n\nComment:\n    ')
HL_RGB = 'FFFFF2CC'   # pale yellow highlight for proposed new rows
COL = {c: i for i, c in enumerate('ABCDEFGHIJKLMNOPQRSTUVWXYZ')}   # letter -> 0-based index


# ---------------------------------------------------------------- shared helpers
def _has_comments(zin):
    return any(re.match(r'xl/(comments|threadedComments/)', n) for n in zin.namelist())


def _threaded_part(guids, refs, texts, dt):
    tc = ''.join(
        f'<threadedComment ref="{r}" dT="{dt}" personId="{PERSON_ID}" id="{guids[i]}">'
        f'<text>{escape(texts[i])}</text></threadedComment>' for i, r in enumerate(refs))
    return ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
            '<ThreadedComments xmlns="http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments" '
            'xmlns:x="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            f'{tc}</ThreadedComments>').encode('utf8')


def _person_part(author):
    return ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
            '<personList xmlns="http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments" '
            'xmlns:x="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            f'<person displayName="{escape(author)}" id="{PERSON_ID}" userId="{escape(author)}" providerId="None"/>'
            '</personList>').encode('utf8')


# ---------------------------------------------------------------- REBUILD mode
def build_rebuild(cfg, source, out):
    import openpyxl
    from openpyxl.comments import Comment
    from openpyxl.styles import PatternFill
    from copy import copy

    sheet = cfg['sheet']
    author = cfg.get('author', 'reviewer')
    dt = cfg.get('dt', '2026-01-01T00:00:00.00')
    trow = int(cfg.get('template_row', 2))
    stage = out + '.stage'

    wb = openpyxl.load_workbook(source)
    ws = wb[sheet]
    fill = PatternFill(start_color='FF' + HL_RGB[2:], end_color='FF' + HL_RGB[2:], fill_type='solid')
    start = ws.max_row + 1
    for i, row in enumerate(cfg.get('new_rows', [])):
        r = start + i
        for c, val in row.items():
            src, tgt = ws[f'{c}{trow}'], ws[f'{c}{r}']
            tgt.value = val
            tgt.font = copy(src.font); tgt.alignment = copy(src.alignment)
            tgt.border = copy(src.border); tgt.number_format = src.number_format
            tgt.fill = fill
    comments = cfg.get('comments', {})
    for coord, text in comments.items():
        cm = Comment(text, author); cm.width, cm.height = 480, 300
        ws[coord].comment = cm
    wb.save(stage)

    zin = zipfile.ZipFile(stage); names = zin.namelist()
    parts = {n: zin.read(n) for n in names}; zin.close()
    cpart = next((n for n in names if re.match(r'xl/comments/comment\d+\.xml$', n)
                  or re.match(r'xl/comments\d+\.xml$', n)), None)
    if cpart is None:
        os.replace(stage, out); return start, len(comments), len(cfg.get('new_rows', []))

    cx = parts[cpart].decode('utf8')
    refs = re.findall(r'<comment ref="([A-Z]+\d+)"', cx)
    guid = {r: '{' + str(uuid.uuid4()).upper() + '}' for r in refs}
    authors = ''.join(f'<author>tc={guid[r]}</author>' for r in refs)
    cx = re.sub(r'<authors>.*?</authors>', f'<authors>{authors}</authors>', cx, count=1, flags=re.S)
    cx = re.sub(r'<comment ref="([A-Z]+\d+)" authorId="\d+"',
                lambda m: f'<comment ref="{m.group(1)}" authorId="{refs.index(m.group(1))}"', cx)
    parts[cpart] = cx.encode('utf8')
    parts['xl/threadedComments/threadedComment1.xml'] = _threaded_part(
        [guid[r] for r in refs], refs, [comments[r] for r in refs], dt)
    parts['xl/persons/person.xml'] = _person_part(author)

    srel = next(n for n in names if re.match(r'xl/worksheets/_rels/.*\.rels$', n))
    parts[srel] = parts[srel].decode('utf8').replace('</Relationships>',
        '<Relationship Id="threadedComment1" '
        'Type="http://schemas.microsoft.com/office/2017/10/relationships/threadedComment" '
        'Target="/xl/threadedComments/threadedComment1.xml"/></Relationships>').encode('utf8')
    wbr = 'xl/_rels/workbook.xml.rels'; wx = parts[wbr].decode('utf8')
    wids = [int(x) for x in re.findall(r'Id="rId(\d+)"', wx)] or [0]
    parts[wbr] = wx.replace('</Relationships>',
        f'<Relationship Id="rId{max(wids)+1}" '
        'Type="http://schemas.microsoft.com/office/2017/10/relationships/person" '
        'Target="/xl/persons/person.xml"/></Relationships>').encode('utf8')
    parts['[Content_Types].xml'] = parts['[Content_Types].xml'].decode('utf8').replace('</Types>',
        '<Override PartName="/xl/threadedComments/threadedComment1.xml" ContentType="application/vnd.ms-excel.threadedcomments+xml"/>'
        '<Override PartName="/xl/persons/person.xml" ContentType="application/vnd.ms-excel.person+xml"/></Types>').encode('utf8')

    if os.path.exists(out): os.remove(out)
    with zipfile.ZipFile(out, 'w', zipfile.ZIP_DEFLATED) as z:
        for n in parts: z.writestr(n, parts[n])
    os.remove(stage)
    return start, len(comments), len(cfg.get('new_rows', []))


# ---------------------------------------------------------------- IN-PLACE mode
def _target_sheet_part(parts, sheet_name):
    wbx = parts['xl/workbook.xml'].decode('utf8')
    m = re.search(r'<sheet[^>]*name="' + re.escape(sheet_name) + r'"[^>]*r:id="(rId\d+)"', wbx) \
        or re.search(r'<sheet[^>]*r:id="(rId\d+)"[^>]*name="' + re.escape(sheet_name) + r'"', wbx)
    if not m:
        sys.exit(f'ERROR: sheet "{sheet_name}" not found in workbook.')
    rid = m.group(1)
    wr = parts['xl/_rels/workbook.xml.rels'].decode('utf8')
    t = re.search(r'Id="' + rid + r'"[^>]*Target="([^"]+)"', wr) \
        or re.search(r'Target="([^"]+)"[^>]*Id="' + rid + r'"', wr)
    tgt = t.group(1).lstrip('/')
    return tgt if tgt.startswith('xl/') else 'xl/' + tgt


def _clone_xf_with_fill(xf, fill_id):
    open_tag = xf[:xf.index('>') + 1]
    rest = xf[xf.index('>') + 1:]
    body = open_tag[:-2] if open_tag.endswith('/>') else open_tag[:-1]
    selfclose = open_tag.endswith('/>')
    body = re.sub(r'\sfillId="\d+"', '', body)
    body = re.sub(r'\sapplyFill="\d+"', '', body)
    body += f' fillId="{fill_id}" applyFill="1"'
    return body + ('/>' if selfclose else '>' + rest)


def _add_highlight_style(parts, tmpl_styles, new_cols):
    """Add a pale-yellow fill + one highlighted xf per distinct template style. Returns {orig_s: hl_s}."""
    st = parts['xl/styles.xml'].decode('utf8')
    fm = re.search(r'<fills count="(\d+)">(.*?)</fills>', st, re.S)
    fills_n = int(fm.group(1)); fill_id = fills_n
    new_fill = (f'<fill><patternFill patternType="solid"><fgColor rgb="{HL_RGB}"/>'
                '<bgColor indexed="64"/></patternFill></fill>')
    st = st.replace(fm.group(0),
                    f'<fills count="{fills_n + 1}">{fm.group(2)}{new_fill}</fills>', 1)
    xm = re.search(r'<cellXfs count="(\d+)">(.*?)</cellXfs>', st, re.S)
    xfs = re.findall(r'<xf\b[^>]*?/>|<xf\b[^>]*?>.*?</xf>', xm.group(2), re.S)
    xfs_n = int(xm.group(1))
    hl = {}; additions = ''
    for si in sorted({tmpl_styles.get(c, 0) for c in new_cols}):
        hl[si] = xfs_n + len(hl)
        additions += _clone_xf_with_fill(xfs[si] if si < len(xfs) else xfs[0], fill_id)
    st = st.replace(xm.group(0),
                    f'<cellXfs count="{xfs_n + len(hl)}">{xm.group(2)}{additions}</cellXfs>', 1)
    parts['xl/styles.xml'] = st.encode('utf8')
    return hl


def add_in_place(cfg, source, out):
    author = cfg.get('author', 'reviewer')
    dt = cfg.get('dt', '2026-01-01T00:00:00.00')
    sheet = cfg['sheet']; trow = int(cfg.get('template_row', 2))
    comments = cfg.get('comments', {}); new_rows = cfg.get('new_rows', [])

    zin = zipfile.ZipFile(source); names = zin.namelist()
    parts = {n: zin.read(n) for n in names}; zin.close()

    sp = _target_sheet_part(parts, sheet)                    # xl/worksheets/sheetN.xml
    srel = f'xl/worksheets/_rels/{os.path.basename(sp)}.rels'
    if srel in parts and re.search(r'relationships/comments"', parts[srel].decode('utf8')):
        sys.exit(f'ERROR: sheet "{sheet}" already has comments. In-place append to an '
                 'already-annotated sheet is unsupported; review a fresh sheet or rebuild.')

    sx = parts[sp].decode('utf8')
    orig_last = max((int(m.group(1)) for m in re.finditer(r'<row r="(\d+)"', sx)), default=0)

    # ---- new rows (inline strings, highlighted) ----
    if new_rows:
        last = orig_last
        tmpl_cells = re.search(rf'<row r="{trow}"[^>]*>(.*?)</row>', sx, re.S)
        tmpl_styles = {}
        if tmpl_cells:
            for cm in re.finditer(r'<c r="([A-Z]+)' + str(trow) + r'"([^>]*)>', tmpl_cells.group(1)):
                sm = re.search(r's="(\d+)"', cm.group(2))
                tmpl_styles[cm.group(1)] = int(sm.group(1)) if sm else 0
        new_cols = sorted({c for row in new_rows for c in row}, key=lambda c: COL[c])
        hl = _add_highlight_style(parts, tmpl_styles, new_cols)
        rows_xml = ''
        for i, row in enumerate(new_rows):
            rn = last + 1 + i
            idxs = sorted((COL[c] for c in row), key=int)
            cells = ''
            for c in sorted(row, key=lambda c: COL[c]):
                s = hl[tmpl_styles.get(c, 0)]
                cells += (f'<c r="{c}{rn}" s="{s}" t="inlineStr"><is>'
                          f'<t xml:space="preserve">{escape(str(row[c]))}</t></is></c>')
            rows_xml += f'<row r="{rn}" spans="{idxs[0]+1}:{idxs[-1]+1}">{cells}</row>'
        sx = sx.replace('</sheetData>', rows_xml + '</sheetData>', 1)
        # extend dimension
        dm = re.search(r'<dimension ref="([A-Z]+)\d+:([A-Z]+)(\d+)"/>', sx)
        if dm:
            sx = sx.replace(dm.group(0),
                            f'<dimension ref="{dm.group(1)}1:{dm.group(2)}{last + len(new_rows)}"/>', 1)

    # ---- comments (threaded) on this sheet ----
    if comments:
        refs = list(comments.keys())
        guids = ['{' + str(uuid.uuid4()).upper() + '}' for _ in refs]
        xruids = ['{' + str(uuid.uuid4()).upper() + '}' for _ in refs]
        used = [int(m) for n in names for m in
                re.findall(r'(?:comments|threadedComment|vmlDrawing)(\d+)', n)]
        K = (max(used) + 1) if used else 1

        clist = ''.join(
            f'<comment ref="{refs[i]}" authorId="{i}" shapeId="0" xr:uid="{xruids[i]}">'
            f'<text><t xml:space="preserve">{escape(BOILER + comments[refs[i]])}</t></text></comment>'
            for i in range(len(refs)))
        parts[f'xl/comments{K}.xml'] = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
            '<comments xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'xmlns:mc="http://schemas.openxmlformats.org/markup-compatibility/2006" mc:Ignorable="xr" '
            'xmlns:xr="http://schemas.microsoft.com/office/spreadsheetml/2014/revision">'
            f'<authors>{"".join(f"<author>tc={g}</author>" for g in guids)}</authors>'
            f'<commentList>{clist}</commentList></comments>').encode('utf8')
        parts[f'xl/threadedComments/threadedComment{K}.xml'] = _threaded_part(
            guids, refs, [comments[r] for r in refs], dt)

        def shape(idx, ref):
            col = COL[re.match(r'([A-Z]+)', ref).group(1)]
            row = int(re.search(r'(\d+)', ref).group(1)) - 1
            return (f'<v:shape id="_x0000_s{K*1024+idx}" type="#_x0000_t202" style=\'position:absolute;'
                    'margin-left:59.25pt;margin-top:1.5pt;width:1107.75pt;height:300pt;z-index:'
                    f'{idx+1};visibility:hidden\' fillcolor="infoBackground [80]" strokecolor="none [81]" '
                    'o:insetmode="auto"><v:fill color2="infoBackground [80]"/><v:shadow color="none [81]" '
                    'obscured="t"/><v:path o:connecttype="none"/><v:textbox style=\'mso-direction-alt:auto\'>'
                    "<div style='text-align:left'></div></v:textbox><x:ClientData ObjectType=\"Note\">"
                    '<x:MoveWithCells/><x:SizeWithCells/><x:Anchor>1, 15, 0, 2, 6, 392, 20, 2</x:Anchor>'
                    f'<x:AutoFill>False</x:AutoFill><x:Row>{row}</x:Row><x:Column>{col}</x:Column>'
                    '</x:ClientData></v:shape>')
        parts[f'xl/drawings/vmlDrawing{K}.vml'] = (
            '<xml xmlns:v="urn:schemas-microsoft-com:vml" xmlns:o="urn:schemas-microsoft-com:office:office" '
            'xmlns:x="urn:schemas-microsoft-com:office:excel"><o:shapelayout v:ext="edit">'
            f'<o:idmap v:ext="edit" data="{K}"/></o:shapelayout>'
            '<v:shapetype id="_x0000_t202" coordsize="21600,21600" o:spt="202" '
            'path="m,l,21600r21600,l21600,xe"><v:stroke joinstyle="miter"/>'
            '<v:path gradientshapeok="t" o:connecttype="rect"/></v:shapetype>'
            f'{"".join(shape(i, refs[i]) for i in range(len(refs)))}</xml>').encode('utf8')

        parts[srel] = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            f'<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/vmlDrawing" Target="../drawings/vmlDrawing{K}.vml"/>'
            f'<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/comments" Target="../comments{K}.xml"/>'
            f'<Relationship Id="rId3" Type="http://schemas.microsoft.com/office/2017/10/relationships/threadedComment" Target="../threadedComments/threadedComment{K}.xml"/>'
            '</Relationships>').encode('utf8') if srel not in parts else parts[srel]
        if '<legacyDrawing' not in sx:
            sx = sx.replace('</worksheet>', '<legacyDrawing r:id="rId1"/></worksheet>', 1)

        ct = parts['[Content_Types].xml'].decode('utf8')
        for frag, mime in ((f'/xl/comments{K}.xml', 'application/vnd.openxmlformats-officedocument.spreadsheetml.comments+xml'),
                           (f'/xl/threadedComments/threadedComment{K}.xml', 'application/vnd.ms-excel.threadedcomments+xml')):
            if frag not in ct:
                ct = ct.replace('</Types>', f'<Override PartName="{frag}" ContentType="{mime}"/></Types>')
        # ensure person + its content-type + workbook relationship
        if 'xl/persons/person.xml' not in parts:
            parts['xl/persons/person.xml'] = _person_part(author)
        elif PERSON_ID not in parts['xl/persons/person.xml'].decode('utf8'):
            pp = parts['xl/persons/person.xml'].decode('utf8')
            parts['xl/persons/person.xml'] = pp.replace('</personList>',
                f'<person displayName="{escape(author)}" id="{PERSON_ID}" userId="{escape(author)}" providerId="None"/></personList>').encode('utf8')
        if '/xl/persons/person.xml' not in ct:
            ct = ct.replace('</Types>', '<Override PartName="/xl/persons/person.xml" ContentType="application/vnd.ms-excel.person+xml"/></Types>')
        parts['[Content_Types].xml'] = ct.encode('utf8')
        wr = parts['xl/_rels/workbook.xml.rels'].decode('utf8')
        if 'relationships/person' not in wr:
            wids = [int(x) for x in re.findall(r'Id="rId(\d+)"', wr)] or [0]
            wr = wr.replace('</Relationships>',
                f'<Relationship Id="rId{max(wids)+1}" '
                'Type="http://schemas.microsoft.com/office/2017/10/relationships/person" '
                'Target="/xl/persons/person.xml"/></Relationships>')
            parts['xl/_rels/workbook.xml.rels'] = wr.encode('utf8')

    parts[sp] = sx.encode('utf8')

    tmp = out + '.tmp'
    with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as z:
        for n, b in parts.items(): z.writestr(n, b)
    os.replace(tmp, out)
    return orig_last + 1, len(comments), len(new_rows)


# ---------------------------------------------------------------- verify
def verify(path):
    import openpyxl, posixpath
    from xml.etree import ElementTree as ET
    z = zipfile.ZipFile(path); names = set(z.namelist())
    for n in names:
        if n.endswith(('.xml', '.rels', '.vml')): ET.fromstring(z.read(n))
    def norm(b, t): return t[1:] if t.startswith('/') else posixpath.normpath(posixpath.join(posixpath.dirname(b), t))
    referenced, dangling = set(), []
    for rp in [n for n in names if n.endswith('.rels')]:
        base = rp.replace('_rels/', '').replace('.rels', '')
        for rel in ET.fromstring(z.read(rp)):
            if rel.get('TargetMode') == 'External': continue
            t = norm(base, rel.get('Target')); referenced.add(t)
            if t not in names: dangling.append((rp, rel.get('Target')))
    orphans = [n for n in names if n not in referenced and not n.endswith('.rels') and n != '[Content_Types].xml']
    ct = z.read('[Content_Types].xml').decode('utf8')
    openpyxl.load_workbook(path)
    ok = not dangling and not orphans and 'threadedcomments+xml' in ct and 'person+xml' in ct
    print('dangling:', dangling or 'none')
    print('orphans :', orphans or 'none')
    print('content-types threaded+person:', 'threadedcomments+xml' in ct and 'person+xml' in ct)
    print('VERIFY:', 'OK' if ok else 'FAIL')
    return ok


if __name__ == '__main__':
    if len(sys.argv) == 3 and sys.argv[1] == '--verify':
        sys.exit(0 if verify(sys.argv[2]) else 1)
    if len(sys.argv) != 2:
        sys.exit(__doc__)
    with open(sys.argv[1], encoding='utf8') as f:
        cfg = json.load(f)
    source = cfg.get('source') or cfg['backup']
    out = cfg.get('out', source)
    if os.path.abspath(source) == os.path.abspath(out) and 'backup' in cfg and 'source' not in cfg:
        sys.exit('ERROR: backup and out must differ in rebuild mode — read clean, write fresh.')
    if os.path.exists(out):                                   # always keep a .PREV backup
        shutil.copyfile(out, out[:-5] + '.PREV.xlsx')
    with zipfile.ZipFile(source) as zt:
        mode = 'in-place' if _has_comments(zt) else 'rebuild'
    start, nc, nr = (add_in_place if mode == 'in-place' else build_rebuild)(cfg, source, out)
    print(f'OK [{mode}] -> {out}  | comments: {nc} | new rows from {start} ({nr})')
    verify(out)
